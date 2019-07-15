#! python3
# blank_row_inserter.py - Inserts M blank rows into the spreadsheet.
# Starting at row N.


import sys
import openpyxl


n, m = 0, 0

if len(sys.argv) >= 3:
    n, m = sys.argv[1], sys.argv[2]


wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

wb_src = openpyxl.load_workbook('produceSales.xlsx')
sheet_src = wb_src.get_sheet_by_name('Sheet')


delta = 0

for row in range(1, sheet_src.max_row+1):

    if row == n:
        delta = m

    for column in range(1, sheet_src.max_column+1):

        cell_src = sheet_src.cell(row=row, column=column)
        cell_new = sheet_new.cell(row=row+delta, column=column)

        cell_new.value = cell_src.value


wb_new.save('produceSalesInsert.xlsx')
print('Done')
