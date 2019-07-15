# python3
# random_chore_assignment_emailer.py - Randomly assigns chores to people from the list


import openpyxl
import random
import smtplib
import sys


wb = openpyxl.load_workbook('chores.xlsx')


# Get chores
chores = []
sheet = wb.get_sheet_by_name('Chores')

for r in range(2, sheet.max_row + 1):  # skip header
    chore = sheet.cell(row=r, column=1).value
    chores.append(chore)


# Get recipients
recipients = {}
sheet = wb.get_sheet_by_name('People')

for r in range(2, sheet.max_row + 1):  # skip header
    name = sheet.cell(row=r, column=1).value
    email = sheet.cell(row=r, column=2).value
    chore = random.choice(chores)
    recipients[name] = {'email': email, 'chore': chore}


# Log in
smtp_obj = smtplib.SMTP_SSL('smtp.gmail.com', 465)
smtp_obj.ehlo()
smtp_obj.login('ivanov.ivan.python.2019', sys.argv[1])


# Send out
for name in recipients.keys():
    email = recipients[name]['email']
    chore = recipients[name]['chore']
    body = "Subject: Chore for today.\n%s!\nYour chore for today is %s" % (name, chore)
    print('Sending email to %s...' % email)
    status = smtp_obj.sendmail('ivanov.ivan.python.2019', email, body)

    if status != {}:
        print('There was a problem sending email to %s: %s' % (email, status))

smtp_obj.quit()
