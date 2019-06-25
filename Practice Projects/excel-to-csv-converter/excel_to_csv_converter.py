#! python3
# excel_to_csv_converter.py - Reads all the Excel files in the current
# working directory and outputs them as CSV files.


import os
import openpyxl
import csv


for excel_file in os.listdir('.'):

    if not excel_file.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)

        csv_filename = excel_file.split('.')[0] + '_' + sheet.title + '.csv'
        csv_file = open(csv_filename, 'w')

        csv_writer = csv.writer(csv_file)

        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell_data = sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_data)

            csv_writer.writerow(row_data)

        csv_file.close()
