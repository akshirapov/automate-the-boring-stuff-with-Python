#! python3
# spreadsheet_cell_inverter.py - Invert row and columns of cells.


import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active


sheet_data = []

for i in range(1, sheet.max_row+1):

    row_data = []

    for j in range(1, sheet.max_column+1):
        cell_data = sheet.cell(row=i, column=j).value
        row_data.append(cell_data)

    sheet_data.append(row_data)


wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

for i in range(len(sheet_data)):

    for j in range(len(sheet_data[i])):
        cell_data = sheet_new.cell(row=j+1, column=i+1).value = sheet_data[i][j]

wb_new.save('example_inverted.xlsx')
print('Done')
