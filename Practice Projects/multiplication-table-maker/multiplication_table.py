#! python3
# multiplication_table_maker.py - Takes a number N and creates
# an N×N multiplication table in an Excel spreadsheet


import sys
import openpyxl
from openpyxl.styles import Font


# Take a number N from command line
size_num = 6
if len(sys.argv) > 1:
    size_num = sys.argv[1]

# Create workbook with sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Set labels with fonts
label_font = Font(bold=True)

for row_num in range(2, size_num+2):
    cell = sheet.cell(row=row_num, column=1)
    cell.font = label_font
    cell.value = row_num-1

for column_num in range(2, size_num+2):
    cell = sheet.cell(row=1, column=column_num)
    cell.font = label_font
    cell.value = column_num-1

# Fill cells
for row_num in range(1, size_num+1):
    for column_num in range(1, size_num+1):
        cell = sheet.cell(row=row_num+1, column=column_num+1)
        cell.value = row_num * column_num

# Save workbook
wb.save('multiplication_table.xlsx')
