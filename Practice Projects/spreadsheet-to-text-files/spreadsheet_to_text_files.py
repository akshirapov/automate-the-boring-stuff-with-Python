#! python3
# spreadsheet_to_text_files.py - Saves text of each column to text file.


import openpyxl

texts = []

wb = openpyxl.load_workbook('spreadsheet.xlsx')
sheet = wb.active


texts = []

for j in range(sheet.max_column):
    text = []
    for i in range(sheet.max_row):
        cell_data = sheet.cell(row=i+1, column=j+1).value
        text.append(cell_data)
    texts.append(text)

for i in range(len(texts)):
    file = open('text_%s.txt' % i, 'w')
    for text_line in texts[i]:
        file.writelines(str(text_line) + '\n')
    file.close()

print('Done.')
